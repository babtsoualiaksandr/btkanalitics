"""Генерация отчётов по Alarm (PDF/XLSX) для отправки в Telegram."""

from __future__ import annotations

import datetime
import io
import logging
from typing import Any, Dict, Iterable, List, Optional, Tuple

from django.conf import settings
from django.db.models import QuerySet
from django.utils import timezone

from agromash.models import Alarm, Monitor, TelegramReportSubscription
from agromash.services.alarm_data_parser import format_alarm_caption, parse_alarm_data
from agromash.va_api_client import VAApiClient


logger = logging.getLogger(__name__)


def _to_aware_dt(value: Optional[int]) -> Optional[datetime.datetime]:
    """BigInteger timestamp -> aware datetime (секунды/миллисекунды)."""
    if value is None:
        return None
    ts = int(value)
    if ts > 1_000_000_000_000:
        ts = ts / 1000.0
    return datetime.datetime.fromtimestamp(ts, tz=datetime.timezone.utc)


def _now_epoch_ms(now: datetime.datetime) -> int:
    return int(now.timestamp() * 1000)


def _dt_to_epoch_ms(dt: datetime.datetime) -> int:
    """aware datetime -> epoch ms."""
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return int(dt.timestamp() * 1000)


def _monitor_ids_int(monitors: Iterable[Monitor]) -> List[int]:
    out: List[int] = []
    for m in monitors:
        try:
            out.append(int(m.monitor_id))
        except Exception:
            continue
    return out


def get_alarms_for_subscription(
    *,
    sub: TelegramReportSubscription,
    now: Optional[datetime.datetime] = None,
) -> QuerySet[Alarm]:
    """QuerySet Alarm для отчёта.

    Диапазон задаётся относительно now:
      start_time between (now - period_from) and (now - period_to)
    """
    now = now or timezone.now()
    now_ms = _now_epoch_ms(now)
    from_ms = now_ms - int(sub.period_from_minutes) * 60_000
    to_ms = now_ms - int(sub.period_to_minutes) * 60_000
    if from_ms > to_ms:
        from_ms, to_ms = to_ms, from_ms

    qs = Alarm.objects.filter(start_time__gte=from_ms, start_time__lte=to_ms)

    # Фильтр по мониторам (если выбраны)
    monitors = list(sub.monitors.all())
    monitor_ids = _monitor_ids_int(monitors)
    if monitor_ids:
        qs = qs.filter(monitor_id__in=monitor_ids)

    return qs.order_by("start_time")


def get_alarms_for_subscription_range(
    *,
    sub: TelegramReportSubscription,
    start: datetime.datetime,
    end: datetime.datetime,
) -> QuerySet[Alarm]:
    """QuerySet Alarm для отчёта по заданному диапазону дат/времени."""

    start_ms = _dt_to_epoch_ms(start)
    end_ms = _dt_to_epoch_ms(end)
    if start_ms > end_ms:
        start_ms, end_ms = end_ms, start_ms

    qs = Alarm.objects.filter(start_time__gte=start_ms, start_time__lte=end_ms)

    monitors = list(sub.monitors.all())
    monitor_ids = _monitor_ids_int(monitors)
    if monitor_ids:
        qs = qs.filter(monitor_id__in=monitor_ids)

    return qs.order_by("start_time")


def build_report_rows(alarms: Iterable[Alarm]) -> List[Dict[str, Any]]:
    """Преобразовать Alarm'ы в строки отчёта."""
    rows: List[Dict[str, Any]] = []
    for a in alarms:
        try:
            parsed = parse_alarm_data(a.data or {})
            summary = format_alarm_caption(parsed)
        except Exception:
            logger.exception("Failed to parse Alarm.data (alarm_id=%s)", a.alarm_id)
            parsed = None
            summary = a.topic

        dt = _to_aware_dt(a.start_time)
        dt_local = timezone.localtime(dt) if dt else None
        rows.append(
            {
                "time": dt_local.strftime("%Y-%m-%d %H:%M:%S") if dt_local else "-",
                "topic": a.topic,
                "monitor_id": a.monitor_id,
                "monitor_name": a.monitor_name,
                "channel_name": (parsed.channel_name if parsed else ""),
                "tags": ", ".join(parsed.tags) if parsed else "",
                "alarm_id": a.alarm_id,
                "event_id": a.event_id,
                # summary: будет добавлен последней колонкой на этапе XLSX/PDF
                "summary": summary,
                # snapshot: будет заполнен при генерации вложений
                "snapshot_url": "",
                "snapshot_bytes": None,
            }
        )
    return rows


def _build_absolute_url(path: str) -> str:
    base = getattr(settings, "BASE_URL", "") or ""
    if not base:
        return path
    if path.startswith("http://") or path.startswith("https://"):
        return path
    if not path.startswith("/"):
        path = "/" + path
    return base.rstrip("/") + path


def _fetch_snapshot_bytes(*, alarm: Alarm, client_cache: Dict[int, VAApiClient]) -> Optional[bytes]:
    """Скачать Alarm.original_quality_snapshot (если возможно) и вернуть bytes.

    ВАЖНО: это может быть дорого по времени/трафику, поэтому при ошибках просто возвращаем None.
    """
    snap = getattr(alarm, "original_quality_snapshot", None)
    if not snap:
        return None
    try:
        account_id = int(alarm.account_id)
    except Exception:
        return None

    client = client_cache.get(account_id)
    if client is None:
        base_url = getattr(settings, "BASE_URL", None)
        if not base_url:
            return None
        client = VAApiClient(account_id=account_id, base_url=base_url)
        client_cache[account_id] = client

    try:
        resp = client.request("GET", snap, timeout=(7.0, 25.0))
        try:
            if resp.status_code != 200:
                return None
            return resp.content
        finally:
            resp.close()
    except Exception:
        return None


def generate_xlsx(*, rows: List[Dict[str, Any]]) -> Optional[bytes]:
    """Сгенерировать XLSX. Возвращает bytes или None, если генератор недоступен."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception:
        logger.warning("openpyxl не установлен — XLSX отчёт не будет сформирован")
        return None

    # Попробуем включить изображения (требуется pillow)
    try:
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage  # type: ignore

        pillow_available = True
    except Exception:
        pillow_available = False
        XLImage = None  # type: ignore
        PILImage = None  # type: ignore

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alarms"

    # split summary на части по разделителю "|" + summary последней колонкой
    max_parts = 0
    for r in rows:
        parts = [p.strip() for p in str(r.get("summary") or "").split("|") if p.strip()]
        max_parts = max(max_parts, len(parts))

    headers = [
        "time",
        "topic",
        "monitor_id",
        "monitor_name",
        "channel_name",
        "tags",
        "alarm_id",
        "event_id",
        "snapshot",  # картинка (если получится)
        "snapshot_url",
    ]
    headers.extend([f"summary_{i+1}" for i in range(max_parts)])
    headers.append("summary")
    ws.append(headers)

    # Чтобы вставлять картинки, нам надо знать координаты строки.
    for row_idx, r in enumerate(rows, start=2):
        summary = str(r.get("summary") or "")
        parts = [p.strip() for p in summary.split("|") if p.strip()]
        summary_cols = parts + [""] * max(0, max_parts - len(parts))

        values = [
            r.get("time", ""),
            r.get("topic", ""),
            r.get("monitor_id", ""),
            r.get("monitor_name", ""),
            r.get("channel_name", ""),
            r.get("tags", ""),
            r.get("alarm_id", ""),
            r.get("event_id", ""),
            "",  # snapshot placeholder
            r.get("snapshot_url", ""),
            *summary_cols,
            summary,
        ]
        ws.append(values)

        # Пытаемся вставить картинку
        snap_bytes = r.get("snapshot_bytes")
        if pillow_available and snap_bytes:
            try:
                img_stream = io.BytesIO(snap_bytes)
                pil_img = PILImage.open(img_stream)
                # уменьшаем до превью
                pil_img.thumbnail((240, 160))
                out = io.BytesIO()
                pil_img.save(out, format="PNG")
                out.seek(0)
                xl_img = XLImage(out)

                # колонка snapshot
                snap_col_idx = headers.index("snapshot") + 1
                cell = f"{get_column_letter(snap_col_idx)}{row_idx}"
                ws.add_image(xl_img, cell)
                ws.row_dimensions[row_idx].height = 80
            except Exception:
                # Если не получилось — оставляем только URL
                pass

    # простая авто-ширина
    for idx, h in enumerate(headers, start=1):
        col = get_column_letter(idx)
        if h == "snapshot":
            ws.column_dimensions[col].width = 18
        else:
            ws.column_dimensions[col].width = max(12, min(60, len(h) + 5))

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _shrink_image_bytes_for_pdf(data: bytes, *, max_dim: int = 400, quality: int = 70) -> bytes:
    """Уменьшить/пережать снимок перед вставкой в PDF.

    В PDF картинка рисуется мелкой (35x25мм), но reportlab.Image вставляет
    исходные байты как есть — при original_quality-снимках с камер (сотни КБ
    каждый) 50 картинок легко раздувают письмо за лимит вложений почты
    (Gmail отклоняет с 552). Пережимаем в JPEG нужного размера; при любой
    ошибке возвращаем исходные байты, чтобы не терять картинку.
    """
    try:
        from PIL import Image as PILImage
    except Exception:
        return data
    try:
        img = PILImage.open(io.BytesIO(data))
        img = img.convert("RGB")
        img.thumbnail((max_dim, max_dim))
        out = io.BytesIO()
        img.save(out, format="JPEG", quality=quality, optimize=True)
        return out.getvalue()
    except Exception:
        return data


def generate_pdf(*, rows: List[Dict[str, Any]], title: str) -> Optional[bytes]:
    """Сгенерировать простой PDF. Возвращает bytes или None, если генератор недоступен."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (  # type: ignore
            Image,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
        from reportlab.lib import colors
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except Exception:
        logger.warning("reportlab не установлен — PDF отчёт не будет сформирован")
        return None

    # Фикс кодировки: пробуем зарегистрировать шрифт с кириллицей.
    # DejaVuSans есть почти на всех Linux; если нет — останется Helvetica.
    font_name = "Helvetica"
    try:
        dejavu_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
        pdfmetrics.registerFont(TTFont("DejaVuSans", dejavu_path))
        font_name = "DejaVuSans"
    except Exception:
        pass

    bio = io.BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=title,
    )

    styles = getSampleStyleSheet()
    style_title = ParagraphStyle(
        name="TitleDejaVu",
        parent=styles["Title"],
        fontName=font_name,
    )
    style_normal = ParagraphStyle(
        name="NormalDejaVu",
        parent=styles["BodyText"],
        fontName=font_name,
        fontSize=9,
        leading=11,
        wordWrap="CJK",  # помогает с длинными строками без пробелов
    )
    style_mono = ParagraphStyle(
        name="MonoDejaVu",
        parent=style_normal,
        fontName=font_name,
        fontSize=8,
        leading=10,
    )

    story = [Paragraph(title, style_title), Spacer(1, 6 * mm)]

    # Формат "чтобы вся инфа была видна": делаем карточку на один Alarm
    # (картинка + многострочный текст с переносами, без обрезки по ширине).
    for r in rows:
        summary = str(r.get("summary") or "")
        parts = [p.strip() for p in summary.split("|") if p.strip()]

        lines: List[str] = []
        lines.append(f"<b>time</b>: {r.get('time', '-')}")
        lines.append(f"<b>topic</b>: {r.get('topic', '')}")
        lines.append(f"<b>monitor</b>: {r.get('monitor_name', '')} (id={r.get('monitor_id', '')})")
        if r.get("channel_name"):
            lines.append(f"<b>channel</b>: {r.get('channel_name', '')}")
        if r.get("tags"):
            lines.append(f"<b>tags</b>: {r.get('tags', '')}")
        if r.get("alarm_id"):
            lines.append(f"<b>alarm_id</b>: {r.get('alarm_id', '')}")
        if r.get("event_id"):
            lines.append(f"<b>event_id</b>: {r.get('event_id', '')}")
        if r.get("snapshot_url"):
            lines.append(f"<b>snapshot_url</b>: {r.get('snapshot_url', '')}")

        if parts:
            for idx, p in enumerate(parts, start=1):
                lines.append(f"<b>summary_{idx}</b>: {p}")
        lines.append(f"<b>summary</b>: {summary}")

        text = Paragraph("<br/>".join(lines), style_normal)

        # Картинка превью слева, если есть
        img_cell = ""
        snap_bytes = r.get("snapshot_bytes")
        if snap_bytes:
            try:
                small_bytes = _shrink_image_bytes_for_pdf(snap_bytes)
                img_cell = Image(io.BytesIO(small_bytes))
                img_cell.drawHeight = 25 * mm
                img_cell.drawWidth = 35 * mm
            except Exception:
                img_cell = ""

        table = Table(
            [[img_cell, text]],
            colWidths=[40 * mm, None],
        )
        table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BOX", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )

        story.append(table)
        story.append(Spacer(1, 4 * mm))

    doc.build(story)
    return bio.getvalue()


def generate_report_attachments(
    *,
    sub: TelegramReportSubscription,
    now: Optional[datetime.datetime] = None,
) -> Tuple[str, List[Tuple[str, bytes, str]], int]:
    """Сформировать файлы отчёта.

    Возвращает:
      - caption
      - список вложений: (filename, content_bytes, mime_type)
      - количество строк (alarms) в отчёте
    """
    now = now or timezone.now()
    alarms_qs = get_alarms_for_subscription(sub=sub, now=now)
    alarms = list(alarms_qs.select_related("account"))
    rows = build_report_rows(alarms)

    # Добавляем ссылки и превью snapshot (best-effort)
    client_cache: Dict[int, VAApiClient] = {}
    # чтобы не раздувать отчёт/время — ограничим количество скачиваемых картинок
    max_images = 50
    img_count = 0
    for r, a in zip(rows, alarms):
        snap = getattr(a, "original_quality_snapshot", None)
        if snap:
            r["snapshot_url"] = _build_absolute_url(str(snap))
        if snap and img_count < max_images:
            b = _fetch_snapshot_bytes(alarm=a, client_cache=client_cache)
            if b:
                r["snapshot_bytes"] = b
                img_count += 1

    caption = (
        f"Отчёт по тревогам: {len(rows)} записей\n"
        f"freq={sub.frequency}, period={sub.period_from_minutes}..{sub.period_to_minutes} мин назад"
    )

    attachments: List[Tuple[str, bytes, str]] = []

    stamp = now.strftime("%Y%m%d_%H%M%S")
    base = f"alarms_report_{stamp}"
    title = f"Alarms report ({stamp})"

    if sub.send_xlsx:
        xlsx = generate_xlsx(rows=rows)
        if xlsx:
            attachments.append((f"{base}.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))

    if sub.send_pdf:
        pdf = generate_pdf(rows=rows, title=title)
        if pdf:
            attachments.append((f"{base}.pdf", pdf, "application/pdf"))

    return caption, attachments, len(rows)


def generate_report_attachments_for_range(
    *,
    sub: TelegramReportSubscription,
    start: datetime.datetime,
    end: datetime.datetime,
    now: Optional[datetime.datetime] = None,
) -> Tuple[str, List[Tuple[str, bytes, str]], int]:
    """Сформировать файлы отчёта по заданному диапазону дат/времени."""

    now = now or timezone.now()
    alarms_qs = get_alarms_for_subscription_range(sub=sub, start=start, end=end)
    alarms = list(alarms_qs.select_related("account"))
    rows = build_report_rows(alarms)

    # Добавляем ссылки и превью snapshot (best-effort)
    client_cache: Dict[int, VAApiClient] = {}
    max_images = 50
    img_count = 0
    for r, a in zip(rows, alarms):
        snap = getattr(a, "original_quality_snapshot", None)
        if snap:
            r["snapshot_url"] = _build_absolute_url(str(snap))
        if snap and img_count < max_images:
            b = _fetch_snapshot_bytes(alarm=a, client_cache=client_cache)
            if b:
                r["snapshot_bytes"] = b
                img_count += 1

    # диапазон отображаем в локальном времени
    if timezone.is_naive(start):
        start = timezone.make_aware(start, timezone.get_current_timezone())
    if timezone.is_naive(end):
        end = timezone.make_aware(end, timezone.get_current_timezone())
    start_local = timezone.localtime(start)
    end_local = timezone.localtime(end)

    caption = (
        f"Отчёт по тревогам: {len(rows)} записей\n"
        f"range={start_local.strftime('%Y-%m-%d %H:%M')}..{end_local.strftime('%Y-%m-%d %H:%M')}"
    )

    attachments: List[Tuple[str, bytes, str]] = []
    stamp = now.strftime("%Y%m%d_%H%M%S")
    base = f"alarms_report_{stamp}"
    title = f"Alarms report ({stamp})"

    if sub.send_xlsx:
        xlsx = generate_xlsx(rows=rows)
        if xlsx:
            attachments.append((f"{base}.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))

    if sub.send_pdf:
        pdf = generate_pdf(rows=rows, title=title)
        if pdf:
            attachments.append((f"{base}.pdf", pdf, "application/pdf"))

    return caption, attachments, len(rows)
