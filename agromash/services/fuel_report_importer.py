from __future__ import annotations

import datetime
import hashlib
import io
import logging
from dataclasses import dataclass
from decimal import Decimal
from typing import IO, Any, Dict, List, Optional, Tuple

from django.contrib.auth.models import User
from django.db import transaction
from django.utils import timezone

from agromash.models import FuelOperation, FuelReport


logger = logging.getLogger(__name__)


class FuelImportError(RuntimeError):
    pass


@dataclass
class FuelImportResult:
    report: FuelReport
    created_rows: int
    skipped_rows: int


EXPECTED_COLUMNS = [
    "Номер топливной карты",
    "Номер подразделения",
    "Дата и время отпуска",
    "Наименование товара/услуги",
    "Код товара/услуги",
    "Количество",
    "Единица измерения",
    "Цена единицы с НДС со скидкой, руб. коп.",
    "Стоимость с НДС со скидкой, руб. коп.",
    "НДС, руб. коп.",
    "Скидка с НДС, руб. коп.",
    "% за услуги",
    "Стоимость услуг с НДС, руб. коп.",
    "Стоимость всего с НДС, руб. коп.",
    "Сумма НДС всего, руб. коп.",
    "Владелец точки обслуживания",
    "Номер точки обслуживания",
    "Номер ТРК/номер секции",
    "Фио водителя",
    "Номер транспортного средства",
]


def _sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _to_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _to_decimal(v: Any) -> Optional[Decimal]:
    if v is None:
        return None
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip().replace(" ", "")
    if not s:
        return None
    # возможна запятая как разделитель
    s = s.replace(",", ".")
    try:
        return Decimal(s)
    except Exception:
        return None


def _to_dt(v: Any) -> Optional[datetime.datetime]:
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        dt = v
    elif isinstance(v, datetime.date):
        dt = datetime.datetime.combine(v, datetime.time.min)
    else:
        s = _to_str(v)
        if not s:
            return None
        # В файле встречается ISO: 2025-12-02 11:58:08
        try:
            dt = datetime.datetime.fromisoformat(s)
        except Exception:
            # fallback: без секунд
            try:
                dt = datetime.datetime.strptime(s, "%Y-%m-%d %H:%M")
            except Exception:
                return None

    # делаем aware (локальная TZ проекта)
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return dt


def _is_total_row(first_cell: str) -> bool:
    s = (first_cell or "").strip().lower()
    return s.startswith("итого")


def _find_header_row(ws) -> Tuple[int, Dict[str, int]]:
    """Возвращает (row_index_1_based, col_map)."""

    # ищем строку, где встречается первая ожидаемая колонка
    target = EXPECTED_COLUMNS[0].lower()
    for r in range(1, min(ws.max_row, 200) + 1):
        values = [(_to_str(ws.cell(row=r, column=c).value)).lower() for c in range(1, ws.max_column + 1)]
        if any(v == target for v in values):
            # строим карту колонок
            col_map: Dict[str, int] = {}
            for c in range(1, ws.max_column + 1):
                v = _to_str(ws.cell(row=r, column=c).value)
                if v:
                    col_map[v] = c

            missing = [h for h in EXPECTED_COLUMNS if h not in col_map]
            if missing:
                raise FuelImportError(f"Header row found but missing columns: {missing}")
            return r, col_map

    raise FuelImportError("Header row not found")


def import_fuel_report_from_xlsx(
    *,
    file_obj: IO[bytes],
    filename: str,
    imported_by: Optional[User],
    period_start: Optional[datetime.date] = None,
    period_end: Optional[datetime.date] = None,
) -> FuelImportResult:
    """Импортирует XLSX в FuelReport + FuelOperation.

    Важно:
      - пропускает строки Итого
      - если "Номер топливной карты" пустой — протягивает значение сверху
    """
    raw = file_obj.read()
    sha = _sha256_bytes(raw)

    # поддержка openpyxl
    try:
        import openpyxl
    except Exception as e:
        raise FuelImportError(f"openpyxl is required to import XLSX: {e}")

    wb = openpyxl.load_workbook(filename=io.BytesIO(raw), data_only=True)
    ws = wb.active

    # meta из первых строк
    title = _to_str(ws.cell(row=1, column=1).value)
    contract_raw = _to_str(ws.cell(row=2, column=1).value)
    org_raw = _to_str(ws.cell(row=3, column=1).value)

    contract_number = contract_raw.replace("По договору №", "").strip()
    organization_name = org_raw.replace("Наименование организации:", "").strip()

    header_row, col_map = _find_header_row(ws)

    last_card = ""
    created = 0
    skipped = 0

    with transaction.atomic():
        report = FuelReport.objects.create(
            title=title,
            contract_number=contract_number,
            organization_name=organization_name,
            period_start=period_start,
            period_end=period_end,
            source_filename=filename,
            source_sha256=sha,
            imported_by=imported_by,
            imported_ok=True,
            import_error="",
        )

        ops: List[FuelOperation] = []
        for r in range(header_row + 1, ws.max_row + 1):
            card = _to_str(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[0]]).value)
            if card:
                # отсекаем строки "Итого" на уровне первой ячейки
                if _is_total_row(card):
                    skipped += 1
                    continue
                last_card = card
            else:
                card = last_card

            # если карта всё ещё пустая — пропускаем (битая строка)
            if not card:
                skipped += 1
                continue

            dt = _to_dt(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[2]]).value)
            product_name = _to_str(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[3]]).value)
            # в некоторых файлах "Итого" может стоять в product_name при пустой карте
            if _is_total_row(product_name):
                skipped += 1
                continue
            if not dt or not product_name:
                skipped += 1
                continue

            ops.append(
                FuelOperation(
                    report=report,
                    card_number=card,
                    department_number=_to_str(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[1]]).value),
                    operation_at=dt,
                    product_name=product_name,
                    product_code=_to_str(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[4]]).value),
                    quantity=_to_decimal(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[5]]).value),
                    unit=_to_str(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[6]]).value),
                    unit_price=_to_decimal(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[7]]).value),
                    cost=_to_decimal(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[8]]).value),
                    vat=_to_decimal(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[9]]).value),
                    discount=_to_decimal(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[10]]).value),
                    service_percent=_to_decimal(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[11]]).value),
                    service_cost=_to_decimal(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[12]]).value),
                    total_cost=_to_decimal(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[13]]).value),
                    total_vat=_to_decimal(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[14]]).value),
                    station_owner=_to_str(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[15]]).value),
                    station_number=_to_str(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[16]]).value),
                    pump_section=_to_str(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[17]]).value),
                    driver_name=_to_str(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[18]]).value),
                    vehicle_number=_to_str(ws.cell(row=r, column=col_map[EXPECTED_COLUMNS[19]]).value),
                )
            )

        if ops:
            FuelOperation.objects.bulk_create(ops, batch_size=1000)
            created = len(ops)

        FuelReport.objects.filter(pk=report.pk).update(rows_count=created)

    return FuelImportResult(report=report, created_rows=created, skipped_rows=skipped)
