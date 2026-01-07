from __future__ import annotations

import datetime
import io
from decimal import Decimal
from typing import Any, Optional

from django.utils import timezone
from django.conf import settings

from agromash.models import Alarm, FuelOperation
from agromash.va_api_client import VAApiClient


# Порядок/набор доступных колонок XLSX.
# `default=True` => включаем по умолчанию (кроме snapshot-колонок).
FUEL_REPORT_XLSX_COLUMNS: list[dict[str, Any]] = [
    {"key": "operation_at", "header": "Дата/время операции", "width": 19, "default": True, "align": "center"},
    {"key": "card_number", "header": "Карта (card_number)", "width": 14, "default": True, "align": "center"},
    {"key": "department_number", "header": "Подразделение", "width": 12, "default": True, "align": "center"},
    {"key": "product_name", "header": "Товар", "width": 20, "default": True, "align": "center"},
    {"key": "product_code", "header": "Код", "width": 10, "default": True, "align": "center"},
    {"key": "quantity", "header": "Количество", "width": 11, "default": True, "align": "center"},
    {"key": "unit", "header": "Ед.", "width": 6, "default": True, "align": "center"},
    {"key": "total_cost", "header": "Стоимость всего", "width": 14, "default": True, "align": "center"},
    {"key": "station_owner", "header": "АЗС (владелец)", "width": 16, "default": True, "align": "center"},
    {"key": "station_number", "header": "АЗС (номер)", "width": 10, "default": True, "align": "center"},
    {"key": "driver_name", "header": "Водитель", "width": 16, "default": True, "align": "center"},
    {"key": "vehicle_number", "header": "ТС", "width": 12, "default": True, "align": "center"},

    {"key": "pi_number", "header": "Номер авто (PlateIdentity)", "width": 14, "default": True, "align": "center"},
    {"key": "pi_state", "header": "Гос.регион", "width": 10, "default": True, "align": "center"},
    {"key": "pi_owner_last", "header": "Владелец (Ф)", "width": 16, "default": True, "align": "center"},
    {"key": "pi_owner_first", "header": "Владелец (И)", "width": 12, "default": True, "align": "center"},
    {"key": "pi_owner_middle", "header": "Владелец (О/card)", "width": 14, "default": True, "align": "center"},
    {"key": "pi_list_name", "header": "Список", "width": 14, "default": True, "align": "center"},
    {"key": "pi_list_level", "header": "Уровень", "width": 8, "default": True, "align": "center"},

    {"key": "fallback_plates", "header": "Fallback plates", "width": 34, "default": True, "align": "leftwrap"},
    {"key": "alarms_count", "header": "Alarm (шт)", "width": 9, "default": True, "align": "center"},
    {"key": "alarms_refs", "header": "Alarm (id@time)", "width": 40, "default": True, "align": "leftwrap"},

    # snapshots (по умолчанию выключены)
    {"key": "snapshot_1", "header": "Snapshot 1", "width": 30, "default": False, "align": "center", "kind": "image"},
    {"key": "snapshot_2", "header": "Snapshot 2", "width": 30, "default": False, "align": "center", "kind": "image"},
    {"key": "snapshot_3", "header": "Snapshot 3", "width": 30, "default": False, "align": "center", "kind": "image"},
    {"key": "snapshot_url_1", "header": "Snapshot URL 1", "width": 44, "default": False, "align": "leftwrap", "kind": "url"},
    {"key": "snapshot_url_2", "header": "Snapshot URL 2", "width": 44, "default": False, "align": "leftwrap", "kind": "url"},
    {"key": "snapshot_url_3", "header": "Snapshot URL 3", "width": 44, "default": False, "align": "leftwrap", "kind": "url"},

    {"key": "analyzed_at", "header": "Проанализировано", "width": 19, "default": True, "align": "center"},
]


def default_fuel_report_xlsx_column_keys() -> list[str]:
    return [c["key"] for c in FUEL_REPORT_XLSX_COLUMNS if c.get("default")]


def _to_local_str(dt: Optional[datetime.datetime]) -> str:
    if not dt:
        return ""
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return timezone.localtime(dt).strftime("%Y-%m-%d %H:%M:%S")


def _parse_iso_datetime(value: Any) -> Optional[datetime.datetime]:
    """Best-effort parse ISO datetime from API payload.

    Supports strings like:
      - 2026-01-01T10:20:30
      - 2026-01-01T10:20:30+03:00
      - 2026-01-01T10:20:30Z
    """
    if not value:
        return None
    if isinstance(value, datetime.datetime):
        return value
    if not isinstance(value, str):
        return None

    v = value.strip()
    if not v:
        return None

    # datetime.fromisoformat doesn't accept trailing "Z".
    if v.endswith("Z"):
        v = f"{v[:-1]}+00:00"

    try:
        return datetime.datetime.fromisoformat(v)
    except ValueError:
        # Fallback: try to parse just the "YYYY-mm-ddTHH:MM:SS" part.
        try:
            return datetime.datetime.fromisoformat(v[:19])
        except ValueError:
            return None


def _to_local_str_from_iso(value: Any) -> str:
    dt = _parse_iso_datetime(value)
    return _to_local_str(dt) if dt else ""


def _format_alarm_ref(alarm: dict[str, Any]) -> str:
    alarm_id = alarm.get("alarm_id") or alarm.get("id") or ""
    started = _to_local_str_from_iso(alarm.get("start_time_iso"))
    return f"{alarm_id}@{started}" if started else f"{alarm_id}@"


def _format_fallback_plate_item(value: Any) -> str:
    """FuelOperation.fallback_plate_numbers item -> human string."""
    if isinstance(value, dict):
        # Выводим только значения (без названий ключей).
        # Сохраняем стабильный порядок через сортировку ключей.
        parts: list[str] = []
        for k in sorted(value.keys(), key=lambda x: str(x)):
            v = value.get(k)
            s = _format_fallback_plate_item(v) if isinstance(v, (dict, list, tuple, set)) else str(v or "")
            if s:
                parts.append(s)
        return " ".join(parts).strip()

    if isinstance(value, (list, tuple, set)):
        parts = [_format_fallback_plate_item(v) for v in value]
        parts = [p for p in parts if p]
        return " ".join(parts).strip()
    return str(value or "")


def _format_fallback_plates(value: Any) -> str:
    items = value if isinstance(value, list) else []
    out = []
    for item in items:
        s = _format_fallback_plate_item(item)
        if s:
            out.append(s)
    return "\n".join(out)


def _fetch_snapshot_bytes(
    *,
    snapshot_path: str,
    account_id: int,
    client_cache: dict[int, VAApiClient],
) -> Optional[bytes]:
    """Скачать snapshot из VA API и вернуть bytes (best-effort).

    Используем VAApiClient (Bearer token в БД), аналогично [`serve_snapshot()`](agromash/views.py:213).
    """
    if not snapshot_path:
        return None

    base_url = getattr(settings, "BASE_URL", None)
    if not base_url:
        return None

    client = client_cache.get(account_id)
    if client is None:
        client = VAApiClient(account_id=int(account_id), base_url=str(base_url))
        client_cache[int(account_id)] = client

    try:
        resp = client.request("GET", str(snapshot_path), timeout=(7.0, 25.0))
        try:
            if resp.status_code != 200:
                return None
            return resp.content
        finally:
            resp.close()
    except Exception:
        return None


def export_fuel_report_to_xlsx_bytes(*, report_id: int, columns: Optional[list[str]] = None) -> bytes:
    """Сформировать XLSX по FuelOperation для указанного FuelReport.

    В выгрузку включаются как исходные поля FuelOperation, так и результаты анализа:
      - PlateIdentity (если проставлен)
      - matched_alarms (если проставлен)
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as e:  # pragma: no cover
        raise RuntimeError("openpyxl is required for XLSX export") from e

    # Попробуем включить изображения (требуется pillow)
    try:
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage  # type: ignore

        pillow_available = True
    except Exception:
        pillow_available = False
        XLImage = None  # type: ignore
        PILImage = None  # type: ignore

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FuelOperations"

    # Какие колонки выводим
    if columns is None:
        selected_keys = default_fuel_report_xlsx_column_keys()
    else:
        want = set(str(k) for k in (columns or []))
        selected_keys = [c["key"] for c in FUEL_REPORT_XLSX_COLUMNS if c.get("key") in want]

    # safety: если пользователь снял все галочки — берём дефолт
    if not selected_keys:
        selected_keys = default_fuel_report_xlsx_column_keys()

    selected_cols = [c for c in FUEL_REPORT_XLSX_COLUMNS if c["key"] in set(selected_keys)]
    headers = [c["header"] for c in selected_cols]

    key_to_col_idx: dict[str, int] = {c["key"]: i + 1 for i, c in enumerate(selected_cols)}
    ALARM_REF_COL_IDX = key_to_col_idx.get("alarms_refs", -1)
    FALLBACK_PLATES_COL_IDX = key_to_col_idx.get("fallback_plates", -1)

    snapshot_keys = ["snapshot_1", "snapshot_2", "snapshot_3"]
    snapshot_url_keys = ["snapshot_url_1", "snapshot_url_2", "snapshot_url_3"]

    need_snapshot_images = any(k in key_to_col_idx for k in snapshot_keys)
    need_snapshot_urls = any(k in key_to_col_idx for k in snapshot_url_keys)
    need_alarms = (
        "alarms_count" in key_to_col_idx
        or "alarms_refs" in key_to_col_idx
        or need_snapshot_images
        or need_snapshot_urls
    )

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(horizontal="center", vertical="center")

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"

    qs = (
        FuelOperation.objects.filter(report_id=report_id)
        .select_related("plate_identity")
        .order_by("operation_at", "id")
    )

    alarm_cache: dict[int, Optional[Alarm]] = {}
    client_cache: dict[int, VAApiClient] = {}
    max_images_total = 200
    images_added = 0

    for op in qs.iterator(chunk_size=2000):
        pi = getattr(op, "plate_identity", None)

        matched = getattr(op, "matched_alarms", None) or []
        alarms_count = len(matched) if need_alarms else 0
        alarms_str = "\n".join(_format_alarm_ref(r) for r in matched) if (need_alarms and matched) else ""

        # Snapshots берём только если их реально выводим.
        snapshot_urls: list[str] = []
        snapshot_bytes_list: list[Optional[bytes]] = []
        if need_snapshot_urls or need_snapshot_images:
            # Берём до 3 снимков по совпавшим Alarm (в том порядке, в каком лежат matched_alarms)
            alarm_pks: list[int] = []
            for r in matched[:3]:
                try:
                    alarm_pks.append(int(r.get("id")))
                except Exception:
                    continue

            for pk in alarm_pks:
                a = alarm_cache.get(pk)
                if a is None and pk not in alarm_cache:
                    try:
                        a = (
                            Alarm.objects.only("id", "account_id", "original_quality_snapshot")
                            .select_related(None)
                            .get(pk=pk)
                        )
                    except Exception:
                        a = None
                    alarm_cache[pk] = a

                if not a or not getattr(a, "original_quality_snapshot", None):
                    snapshot_urls.append("")
                    snapshot_bytes_list.append(None)
                    continue

                snap_path = str(a.original_quality_snapshot)
                snapshot_urls.append(snap_path)

                if not need_snapshot_images or not pillow_available or images_added >= max_images_total:
                    snapshot_bytes_list.append(None)
                    continue

                b = _fetch_snapshot_bytes(
                    snapshot_path=snap_path,
                    account_id=int(a.account_id),
                    client_cache=client_cache,
                )
                snapshot_bytes_list.append(b)

        # Собираем значения по ключам
        values: dict[str, Any] = {
            "operation_at": _to_local_str(op.operation_at),
            "card_number": str(op.card_number or ""),
            "department_number": str(op.department_number or ""),
            "product_name": str(op.product_name or ""),
            "product_code": str(op.product_code or ""),
            "quantity": float(op.quantity) if isinstance(op.quantity, Decimal) else (op.quantity or ""),
            "unit": str(op.unit or ""),
            "total_cost": float(op.total_cost) if isinstance(op.total_cost, Decimal) else (op.total_cost or ""),
            "station_owner": str(op.station_owner or ""),
            "station_number": str(op.station_number or ""),
            "driver_name": str(op.driver_name or ""),
            "vehicle_number": str(op.vehicle_number or ""),
            "pi_number": str(getattr(pi, "number", "") or ""),
            "pi_state": str(getattr(pi, "state", "") or ""),
            "pi_owner_last": str(getattr(pi, "owner_last_name", "") or ""),
            "pi_owner_first": str(getattr(pi, "owner_first_name", "") or ""),
            "pi_owner_middle": str(getattr(pi, "owner_middle_name", "") or ""),
            "pi_list_name": str(getattr(pi, "list_name", "") or ""),
            "pi_list_level": getattr(pi, "list_level", "") if pi else "",
            "fallback_plates": _format_fallback_plates(getattr(op, "fallback_plate_numbers", None)),
            "alarms_count": alarms_count,
            "alarms_refs": alarms_str,
            "snapshot_1": "",
            "snapshot_2": "",
            "snapshot_3": "",
            "snapshot_url_1": snapshot_urls[0] if len(snapshot_urls) > 0 else "",
            "snapshot_url_2": snapshot_urls[1] if len(snapshot_urls) > 1 else "",
            "snapshot_url_3": snapshot_urls[2] if len(snapshot_urls) > 2 else "",
            "analyzed_at": _to_local_str(getattr(op, "analyzed_at", None)),
        }

        row = [values.get(k, "") for k in selected_keys]
        ws.append(row)

        # Выравнивание строк
        row_idx = ws.max_row
        for c in selected_cols:
            col_idx = key_to_col_idx.get(c["key"])
            if not col_idx:
                continue
            align = c.get("align")
            if align == "leftwrap":
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                ws.cell(row=row_idx, column=col_idx).alignment = data_align

        # Вставляем превьюшки снимков (best-effort)
        if need_snapshot_images and pillow_available and images_added < max_images_total:
            for i, snap_bytes in enumerate(snapshot_bytes_list[:3]):
                key = f"snapshot_{i+1}"
                col_idx = key_to_col_idx.get(key)
                if not col_idx:
                    continue
                if not snap_bytes or images_added >= max_images_total:
                    continue
                try:
                    img_stream = io.BytesIO(snap_bytes)
                    pil_img = PILImage.open(img_stream)
                    pil_img.thumbnail((240, 160))
                    out = io.BytesIO()
                    pil_img.save(out, format="PNG")
                    out.seek(0)

                    xl_img = XLImage(out)
                    cell = f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}"
                    ws.add_image(xl_img, cell)
                    ws.row_dimensions[row_idx].height = 80
                    images_added += 1
                except Exception:
                    continue

    # ширины колонок
    for key, col_idx in key_to_col_idx.items():
        col = next((c for c in selected_cols if c["key"] == key), None)
        if not col:
            continue
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = float(col.get("width") or 12)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
