from __future__ import annotations

import datetime
import io
import json
import mimetypes
import logging
from typing import Dict, Iterable, List, Optional, Tuple

from django.conf import settings
from django import forms
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import FileResponse, Http404, HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_GET, require_http_methods, require_POST

from agromash.services.common import (
    ALARM_EPOCH_MS_THRESHOLD,
    alarm_epoch_to_aware_dt,
    assert_events_access,
)

from .models import Alarm, AlarmCase, AlarmDocument, Monitor, UserMonitorAccess
from .va_api_client import VAApiClient


logger = logging.getLogger(__name__)


_assert_events_access = assert_events_access
_alarm_ts_to_aware_utc = alarm_epoch_to_aware_dt


def _json_dump(v) -> str:
    try:
        return json.dumps(v, ensure_ascii=False, sort_keys=True)
    except Exception:
        return str(v)


def _identities_to_text(raw) -> str:
    """Преобразовать plate_identities/face_identities в человекочитаемый текст.

    Требование: в экспорте (Excel/PDF) не показывать JSON, а выводить значения
    объектов без названий ключей.

    Вход обычно выглядит как список объектов вида:
    [{"list": {"id": 104, "name": "БУЭС", "level": 1}, "plates": [{...}]}]
    """

    def _clean_str(v) -> str:
        s = str(v or '').strip()
        return s

    def _join_parts(parts: Iterable[str], sep: str = ' ') -> str:
        out: List[str] = []
        for p in parts:
            s = _clean_str(p)
            if s and s not in out:
                out.append(s)
        return sep.join(out)

    def _fmt_list_meta(v) -> str:
        if not isinstance(v, dict):
            return _clean_str(v)
        name = _clean_str(v.get('name'))
        level = _clean_str(v.get('level'))
        if name and level:
            return f"{name} (ур.{level})"
        return name or level or ''

    def _fmt_plate(v) -> str:
        if isinstance(v, dict):
            return _join_parts(
                [
                    v.get('state'),
                    v.get('number') or v.get('plate') or v.get('plate_number'),
                    v.get('owner_last_name'),
                    v.get('owner_first_name'),
                    v.get('owner_middle_name'),
                ]
            )
        return _clean_str(v)

    def _flatten_values(v) -> List[str]:
        # Универсальный fallback: собрать все скалярные значения в читабельный список.
        if v is None:
            return []
        if isinstance(v, (str, int, float, bool)):
            s = _clean_str(v)
            return [s] if s else []
        if isinstance(v, (list, tuple)):
            out: List[str] = []
            for x in v:
                out.extend(_flatten_values(x))
            return out
        if isinstance(v, dict):
            # Стабильный порядок: сначала наиболее полезные поля.
            preferred = [
                'name',
                'level',
                'state',
                'number',
                'plate',
                'plate_number',
                'owner_last_name',
                'owner_first_name',
                'owner_middle_name',
                'full_name',
                'last_name',
                'first_name',
                'middle_name',
                'title',
                'value',
                'id',
            ]
            keys = [k for k in preferred if k in v] + [k for k in sorted(v.keys()) if k not in preferred]
            out: List[str] = []
            for k in keys:
                out.extend(_flatten_values(v.get(k)))
            return out
        # datetime / прочие типы
        s = _clean_str(v)
        return [s] if s else []

    if not raw:
        return ''

    # Нормализуем в список "элементов".
    if isinstance(raw, list):
        items = raw
    elif isinstance(raw, dict):
        items = [raw]
    else:
        return _clean_str(raw)

    lines: List[str] = []
    for item in items:
        if not item:
            continue

        # Основной ожидаемый формат: {list: {...}, plates: [...]}
        if isinstance(item, dict):
            list_part = _fmt_list_meta(item.get('list'))

            plates = item.get('plates')
            if isinstance(plates, list):
                plates_part = '; '.join([p for p in (_fmt_plate(x) for x in plates) if p])
                line = ' — '.join([x for x in [list_part, plates_part] if x])
                if line:
                    lines.append(line)
                    continue

            # Более общий формат: list + прочие значения без ключей.
            rest = {k: v for k, v in item.items() if k != 'list'}
            rest_part = _join_parts(_flatten_values(rest), sep=' ')
            line = ' — '.join([x for x in [list_part, rest_part] if x])
            if line:
                lines.append(line)
                continue

        # Fallback
        flat = _join_parts(_flatten_values(item), sep=' ')
        if flat:
            lines.append(flat)

    return "\n".join(lines)


def _plate_numbers_from_alarm(a: Alarm) -> str:
    """Список номеров из Alarm.

    Источники (в порядке приоритета):
    1. Alarm.plate_identities (для PlateMatched и других)
    2. Alarm.data.params.plate.number (для PlateNotMatched)
    """

    nums: List[str] = []

    def _add(x) -> None:
        s = str(x or '').strip()
        if s and s not in nums:
            nums.append(s)

    # 1. Пробуем plate_identities
    raw = getattr(a, 'plate_identities', None)
    if raw:
        def _walk(v, *, allow_scalar: bool = False) -> None:
            if v is None:
                return
            if isinstance(v, (str, int, float, bool)):
                if allow_scalar:
                    _add(v)
                return
            if isinstance(v, (list, tuple)):
                for x in v:
                    _walk(x, allow_scalar=True)
                return
            if isinstance(v, dict):
                for k in ('number', 'plate', 'plate_number'):
                    if k in v:
                        _add(v.get(k))
                for x in v.values():
                    _walk(x, allow_scalar=False)
                return
            if allow_scalar:
                _add(v)

        try:
            _walk(raw, allow_scalar=True)
        except Exception:
            pass

    # 2. Если номер не найден и topic=PlateNotMatched, пробуем data.params.plate
    if not nums and str(getattr(a, 'topic', '') or '') == 'PlateNotMatched':
        data = getattr(a, 'data', None)
        if isinstance(data, dict):
            params = data.get('params')
            if isinstance(params, dict):
                plate = params.get('plate')
                if isinstance(plate, dict):
                    plate_number = plate.get('number')
                    if plate_number:
                        _add(plate_number)

    return ', '.join(nums)


def _epoch_bounds(dt: datetime.datetime) -> Tuple[int, int]:
    """aware dt -> (epoch_sec, epoch_ms)."""
    dt_utc = dt.astimezone(datetime.timezone.utc)
    sec = int(dt_utc.timestamp())
    return sec, sec * 1000


def _alarm_start_time_range_q(*, dt_from: Optional[datetime.datetime], dt_to: Optional[datetime.datetime]) -> Q:
    """Фильтр по Alarm.start_time, учитывая что start_time может быть в секундах или миллисекундах.

    Идея:
      - сек: значения < 1e12
      - мс: значения >= 1e12
    """

    TH_MS = ALARM_EPOCH_MS_THRESHOLD

    # prepare bounds
    from_sec = from_ms = None
    to_sec = to_ms = None
    if dt_from:
        from_sec, from_ms = _epoch_bounds(dt_from)
    if dt_to:
        to_sec, to_ms = _epoch_bounds(dt_to)

    q_sec = Q(start_time__lt=TH_MS)
    q_ms = Q(start_time__gte=TH_MS)

    if from_sec is not None:
        q_sec &= Q(start_time__gte=int(from_sec))
        q_ms &= Q(start_time__gte=int(from_ms))
    if to_sec is not None:
        q_sec &= Q(start_time__lte=int(to_sec))
        q_ms &= Q(start_time__lte=int(to_ms))

    return q_sec | q_ms


def _parse_local_dt(value: str) -> Optional[datetime.datetime]:
    s = (value or "").strip()
    if not s:
        return None
    try:
        # datetime-local: 'YYYY-MM-DDTHH:MM' or with seconds
        dt = datetime.datetime.fromisoformat(s)
    except Exception:
        return None

    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return dt


def _allowed_monitors_qs(user) -> Iterable[Monitor]:
    return (
        Monitor.objects.filter(user_accesses__user=user, user_accesses__enabled=True)
        .distinct()
        .order_by('monitor_id')
    )


def _allowed_monitor_id_ints(monitors: Iterable[Monitor]) -> List[int]:
    res: List[int] = []
    for m in monitors:
        try:
            res.append(int(m.monitor_id))
        except Exception:
            continue
    return res


def _monitor_filter_q(*, monitors: Iterable[Monitor]) -> Optional[Q]:
    """Построить Q для Alarm по набору Monitor.

    Учитываем, что:
    - исторически `Alarm.monitor_id` хранится как int (но это не FK)
    - новый FK `Alarm.monitor_ref` может быть NULL для части записей
    - `Monitor.monitor_id` — строка и может быть нечисловой
    """

    mons = list(monitors)
    pks = [m.pk for m in mons if getattr(m, 'pk', None)]
    mids_int = _allowed_monitor_id_ints(mons)

    q = Q()
    if pks:
        q |= Q(monitor_ref_id__in=pks)
    if mids_int:
        q |= Q(monitor_id__in=mids_int)
    # Если не удалось собрать ни одного условия — вернём None,
    # чтобы вызывающий код не применял пустой Q() (который даст "всё").
    if not pks and not mids_int:
        return None
    return q


def _user_can_access_alarm(*, user, alarm: Alarm) -> bool:
    if getattr(user, 'is_staff', False) or getattr(user, 'is_superuser', False):
        return True

    monitor_pk = getattr(alarm, 'monitor_ref_id', None)
    if monitor_pk:
        return UserMonitorAccess.objects.filter(user=user, monitor_id=int(monitor_pk), enabled=True).exists()

    # best-effort: match via numeric monitor_id -> Monitor.monitor_id
    try:
        m = Monitor.objects.filter(monitor_id=str(int(alarm.monitor_id))).only('id').first()
    except Exception:
        m = None
    if not m:
        return False
    return UserMonitorAccess.objects.filter(user=user, monitor_id=int(m.id), enabled=True).exists()


class AlarmFilterForm(forms.Form):
    monitors = forms.ModelMultipleChoiceField(
        queryset=Monitor.objects.none(),
        required=False,
        widget=forms.CheckboxSelectMultiple(attrs={'class': 'btk-checkboxes'}),
        label='Мониторы',
    )
    date_from = forms.DateTimeField(
        required=False,
        label='С',
        input_formats=["%Y-%m-%dT%H:%M", "%Y-%m-%dT%H:%M:%S"],
        widget=forms.DateTimeInput(attrs={'type': 'datetime-local'}),
    )
    date_to = forms.DateTimeField(
        required=False,
        label='По',
        input_formats=["%Y-%m-%dT%H:%M", "%Y-%m-%dT%H:%M:%S"],
        widget=forms.DateTimeInput(attrs={'type': 'datetime-local'}),
    )

    limit = forms.ChoiceField(
        required=False,
        label='Лимит',
        choices=(
            ('10', '10'),
            ('20', '20'),
            ('50', '50'),
            ('100', '100'),
            ('200', '200'),
            ('500', '500'),
            ('1000', '1000'),
        ),
        initial='20',
    )

    def __init__(self, *args, allowed_monitors_qs=None, **kwargs):
        super().__init__(*args, **kwargs)
        if allowed_monitors_qs is not None:
            self.fields['monitors'].queryset = allowed_monitors_qs


def _filter_alarms(*, user, cleaned: dict, sort: str = 'time', direction: str = 'desc') -> List[Alarm]:
    allowed_monitors = list(_allowed_monitors_qs(user))
    selected_monitors = list(cleaned.get('monitors') or [])

    # Логика:
    # - если не выбран ни один монитор — не показываем события
    # - если выбраны мониторы — фильтруем по ним
    if not selected_monitors:
        return []

    effective_monitors = selected_monitors
    force_filter = True

    if force_filter and not effective_monitors:
        return []

    qs = (
        Alarm.objects.all()
        .select_related('account', 'monitor_ref')
        .prefetch_related('case', 'case__documents')
    )

    if force_filter:
        # Фильтрация по выбранным (или разрешённым) мониторам через monitor_ref и/или monitor_id.
        q_mon = _monitor_filter_q(monitors=effective_monitors)
        if q_mon is None:
            return []
        qs = qs.filter(q_mon)

    # Даты (корректный фильтр для сек/мс)
    dt_from = cleaned.get('date_from')
    dt_to = cleaned.get('date_to')
    if dt_from or dt_to:
        qs = qs.filter(_alarm_start_time_range_q(dt_from=dt_from, dt_to=dt_to))

    try:
        limit = int(cleaned.get('limit') or 200)
    except Exception:
        limit = 200
    limit = max(10, min(limit, 1000))

    # --- Sorting ---
    sort = str(sort or 'time')
    direction = str(direction or 'desc').lower()
    if direction not in ('asc', 'desc'):
        direction = 'desc'
    reverse = direction == 'desc'

    # DB-friendly sorts first (до лимита)
    if sort == 'monitor':
        qs = qs.order_by(('-' if reverse else '') + 'monitor_name', '-start_time')
    elif sort == 'topic':
        qs = qs.order_by(('-' if reverse else '') + 'topic', '-start_time')
    elif sort == 'time':
        qs = qs.order_by(('-' if reverse else '') + 'start_time')
    else:
        # fallback
        qs = qs.order_by('-start_time')

    alarms = list(qs[:limit])

    # Localized display fields
    for a in alarms:
        a.start_dt_local = timezone.localtime(_alarm_ts_to_aware_utc(a.start_time)) if a.start_time else None
        a.end_dt_local = timezone.localtime(_alarm_ts_to_aware_utc(a.end_time)) if a.end_time else None
        # Номера автомобилей (для всех событий, не только PlateMatched)
        a.plate_numbers = _plate_numbers_from_alarm(a)
        # Описание из карточки события
        case = getattr(a, 'case', None)
        a.case_description = str(getattr(case, 'description', '') or '') if case else ''

    # Python-level sort для вычисляемых/не-БД полей
    if sort == 'plate':
        alarms.sort(key=lambda a: (getattr(a, 'plate_numbers', '') or ''), reverse=reverse)
    elif sort == 'description':
        alarms.sort(key=lambda a: (getattr(a, 'case_description', '') or ''), reverse=reverse)
    return alarms

@login_required
@require_GET
def events_list(request: HttpRequest):
    _assert_events_access(request.user)
    user = request.user
    allowed_monitors = _allowed_monitors_qs(user)

    # По умолчанию: сегодня с 00:00 до конца дня.
    # В UI это обычно ожидается как "текущая дата" и условные "24:00".
    now = timezone.localtime(timezone.now())
    start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_day = now.replace(hour=23, minute=59, second=0, microsecond=0)
    initial = {
        'date_from': start_of_day.strftime('%Y-%m-%dT%H:%M'),
        'date_to': end_of_day.strftime('%Y-%m-%dT%H:%M'),
        'limit': '20',
    }

    # Сортировка таблицы
    sort = str(request.GET.get('sort') or 'time')
    direction = str(request.GET.get('dir') or 'desc').lower()
    allowed_sorts = {'time', 'monitor', 'topic', 'plate', 'description'}
    if sort not in allowed_sorts:
        sort = 'time'
    if direction not in ('asc', 'desc'):
        direction = 'desc'

    def _build_sort_url(key: str) -> str:
        q = request.GET.copy()
        # toggle если кликаем по активному столбцу
        if key == sort:
            next_dir = 'desc' if direction == 'asc' else 'asc'
        else:
            next_dir = 'asc'
        q['sort'] = key
        q['dir'] = next_dir
        enc = q.urlencode()
        return f"?{enc}" if enc else "?"

    sort_urls = {
        'monitor': _build_sort_url('monitor'),
        'topic': _build_sort_url('topic'),
        'plate': _build_sort_url('plate'),
        'time': _build_sort_url('time'),
        'description': _build_sort_url('description'),
    }

    form = AlarmFilterForm(request.GET or None, allowed_monitors_qs=allowed_monitors, initial=initial)
    if form.is_valid():
        alarms = _filter_alarms(user=user, cleaned=form.cleaned_data, sort=sort, direction=direction)
        try:
            limit_value = int(form.cleaned_data.get('limit') or 0)
        except Exception:
            limit_value = 0
    else:
        alarms = []
        limit_value = int(initial.get('limit') or 20)

    # Поля для экспорта: все поля Alarm + вложенные модели
    export_columns = _export_columns(request)
    export_fields = [(k, export_columns[k][0]) for k in sorted(export_columns.keys())]
    export_defaults = {
        'alarm_id',
        'topic',
        'monitor_name',
        'start_time_local',
        'end_time_local',
        'event_id',
        'plate_numbers',
        'case_description',
        'case_note',
        'documents_count',
        'documents_urls',
    }

    return render(
        request,
        'agromash/events_list.html',
        {
            'form': form,
            'alarms': alarms,
            'limit_value': limit_value,
            'sort': sort,
            'dir': direction,
            'sort_urls': sort_urls,
            'selected_monitor_pks': request.GET.getlist('monitors'),
            'date_from_raw': str(request.GET.get('date_from') or ''),
            'date_to_raw': str(request.GET.get('date_to') or ''),
            'limit_raw': str(request.GET.get('limit') or ''),
            'export_fields': export_fields,
            'export_defaults': export_defaults,
        },
    )


@login_required
@require_GET
def events_table_body(request: HttpRequest):
    """HTML-фрагмент строк таблицы событий (tbody).

    Используется фронтендом для автообновления таблицы без перезагрузки страницы.
    """

    _assert_events_access(request.user)
    user = request.user
    allowed_monitors = _allowed_monitors_qs(user)

    # По умолчанию: сегодня с 00:00 до конца дня (как в events_list)
    now = timezone.localtime(timezone.now())
    start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_day = now.replace(hour=23, minute=59, second=0, microsecond=0)
    initial = {
        'date_from': start_of_day.strftime('%Y-%m-%dT%H:%M'),
        'date_to': end_of_day.strftime('%Y-%m-%dT%H:%M'),
        'limit': '20',
    }

    sort = str(request.GET.get('sort') or 'time')
    direction = str(request.GET.get('dir') or 'desc').lower()
    allowed_sorts = {'time', 'monitor', 'topic', 'plate', 'description'}
    if sort not in allowed_sorts:
        sort = 'time'
    if direction not in ('asc', 'desc'):
        direction = 'desc'

    form = AlarmFilterForm(request.GET or None, allowed_monitors_qs=allowed_monitors, initial=initial)
    if form.is_valid():
        alarms = _filter_alarms(user=user, cleaned=form.cleaned_data, sort=sort, direction=direction)
    else:
        alarms = []

    return render(request, 'agromash/_events_table_body.html', {'alarms': alarms})


def _render_case_partial(request: HttpRequest, *, alarm: Alarm, case: AlarmCase) -> HttpResponse:
    docs = list(case.documents.all().order_by('-uploaded_at'))
    return render(
        request,
        'agromash/alarm_case_modal.html',
        {
            'alarm': alarm,
            'case': case,
            'documents': docs,
        },
    )

@login_required
@require_http_methods(["GET", "POST"])
def alarm_case_modal(request: HttpRequest, alarm_pk: int):
    _assert_events_access(request.user)
    alarm = get_object_or_404(Alarm, pk=int(alarm_pk))
    if not _user_can_access_alarm(user=request.user, alarm=alarm):
        raise Http404('Forbidden')

    case, created = AlarmCase.objects.get_or_create(
        alarm=alarm,
        defaults={
            'created_by': request.user,
            'updated_by': request.user,
        },
    )

    if request.method == 'POST':
        case.description = str(request.POST.get('description') or '').strip()
        case.note = str(request.POST.get('note') or '').strip()
        case.updated_by = request.user
        case.save(update_fields=['description', 'note', 'updated_by', 'updated_at'])

        files = request.FILES.getlist('documents')
        for f in files:
            if not f:
                continue
            AlarmDocument.objects.create(
                case=case,
                file=f,
                title=str(getattr(f, 'name', '') or ''),
                uploaded_by=request.user,
            )

    return _render_case_partial(request, alarm=alarm, case=case)

@login_required
@require_POST
def alarm_document_delete(request: HttpRequest, doc_pk: int):
    _assert_events_access(request.user)
    doc = get_object_or_404(AlarmDocument.objects.select_related('case', 'case__alarm'), pk=int(doc_pk))
    alarm = doc.case.alarm
    if not _user_can_access_alarm(user=request.user, alarm=alarm):
        raise Http404('Forbidden')

    case = doc.case
    doc.delete()
    return _render_case_partial(request, alarm=alarm, case=case)

@login_required
@require_GET
def alarm_document_file(request: HttpRequest, doc_pk: int):
    _assert_events_access(request.user)
    doc = get_object_or_404(AlarmDocument.objects.select_related('case', 'case__alarm'), pk=int(doc_pk))
    alarm = doc.case.alarm
    if not _user_can_access_alarm(user=request.user, alarm=alarm):
        raise Http404('Forbidden')

    if not getattr(doc, 'file', None):
        raise Http404('File missing')

    path = doc.file.path
    content_type, _ = mimetypes.guess_type(path)
    resp = FileResponse(open(path, 'rb'), content_type=content_type or 'application/octet-stream')
    # по умолчанию отображаем inline; при необходимости можно добавить ?download=1
    if str(request.GET.get('download') or '').strip() == '1':
        resp['Content-Disposition'] = f'attachment; filename="{doc.file.name.split("/")[-1]}"'
    return resp


def _export_columns(request: Optional[HttpRequest] = None) -> Dict[str, Tuple[str, callable]]:
    # --- базовые поля Alarm (все поля модели) ---
    def _field(name: str):
        return lambda a: _json_dump(getattr(a, name, None))

    def _identities_field(name: str):
        return lambda a: _identities_to_text(getattr(a, name, None))

    def _start_local(a: Alarm) -> str:
        dt = _alarm_ts_to_aware_utc(a.start_time)
        return timezone.localtime(dt).strftime('%Y-%m-%d %H:%M:%S') if dt else ''

    def _end_local(a: Alarm) -> str:
        dt = _alarm_ts_to_aware_utc(a.end_time)
        return timezone.localtime(dt).strftime('%Y-%m-%d %H:%M:%S') if dt else ''

    def _desc(a: Alarm) -> str:
        case = getattr(a, 'case', None)
        return str(getattr(case, 'description', '') or '') if case else ''

    def _note(a: Alarm) -> str:
        case = getattr(a, 'case', None)
        return str(getattr(case, 'note', '') or '') if case else ''

    def _docs(a: Alarm) -> List[AlarmDocument]:
        case = getattr(a, 'case', None)
        if not case:
            return []
        try:
            return list(case.documents.all().order_by('-uploaded_at'))
        except Exception:
            return []

    def _docs_count(a: Alarm) -> str:
        return str(len(_docs(a)))

    def _docs_titles(a: Alarm) -> str:
        docs = _docs(a)
        return "\n".join([
            (str(d.title or '').strip() or str(getattr(d.file, 'name', '') or '').split('/')[-1])
            for d in docs
        ])

    def _docs_urls(a: Alarm) -> str:
        docs = _docs(a)
        rel = [f"/events/docs/{d.pk}/file/" for d in docs]
        if request is not None:
            return "\n".join([request.build_absolute_uri(x) for x in rel])
        return "\n".join(rel)

    def _plate_numbers(a: Alarm) -> str:
        return _plate_numbers_from_alarm(a)

    def _case_field(field_name: str) -> str:
        def _get(a: Alarm) -> str:
            case = getattr(a, 'case', None)
            return _json_dump(getattr(case, field_name, None)) if case else ''

        return _get

    def _case_user(field_name: str) -> str:
        def _get(a: Alarm) -> str:
            case = getattr(a, 'case', None)
            u = getattr(case, field_name, None) if case else None
            return str(getattr(u, 'username', '') or getattr(u, 'email', '') or getattr(u, 'id', '') or '')

        return _get

    def _account_field(field_name: str) -> str:
        def _get(a: Alarm) -> str:
            acc = getattr(a, 'account', None)
            return str(getattr(acc, field_name, '') or '') if acc else ''

        return _get

    def _monitor_ref_field(field_name: str) -> str:
        def _get(a: Alarm) -> str:
            m = getattr(a, 'monitor_ref', None)
            return str(getattr(m, field_name, '') or '') if m else ''

        return _get

    # Поля Alarm (все)
    columns: Dict[str, Tuple[str, callable]] = {
        'id': ('id', lambda a: str(getattr(a, 'id', '') or '')),
        'monitor_id': ('monitor_id', lambda a: str(getattr(a, 'monitor_id', '') or '')),
        'monitor_name': ('monitor_name', lambda a: str(getattr(a, 'monitor_name', '') or '')),
        'monitor_ref_id': ('monitor_ref_id', lambda a: str(getattr(a, 'monitor_ref_id', '') or '')),
        'alarm_id': ('alarm_id', lambda a: str(getattr(a, 'alarm_id', '') or '')),
        'topic': ('topic', lambda a: str(getattr(a, 'topic', '') or '')),
        'start_time': ('start_time', lambda a: str(getattr(a, 'start_time', '') or '')),
        'end_time': ('end_time', lambda a: str(getattr(a, 'end_time', '') or '')),
        'event_id': ('event_id', lambda a: str(getattr(a, 'event_id', '') or '')),
        'original_quality_snapshot': ('original_quality_snapshot', _field('original_quality_snapshot')),
        'plate_identities': ('plate_identities', _identities_field('plate_identities')),
        'face_identities': ('face_identities', _identities_field('face_identities')),
        'snapshots': ('snapshots', _field('snapshots')),
        'data': ('data', _field('data')),
        'account_id': ('account_id', lambda a: str(getattr(a, 'account_id', '') or '')),
    }

    # Удобные производные поля
    columns.update(
        {
            'start_time_local': ('start_time_local', _start_local),
            'end_time_local': ('end_time_local', _end_local),
            'plate_numbers': ('plate_numbers', _plate_numbers),
        }
    )

    # Вложенная модель аккаунта
    columns.update(
        {
            'account_name': ('account.name', _account_field('name')),
            'account_contract': ('account.contract', _account_field('contract')),
            'account_organization': ('account.organization', _account_field('organization')),
        }
    )

    # Вложенная модель Monitor (FK)
    columns.update(
        {
            'monitor_ref_monitor_id': ('monitor_ref.monitor_id', _monitor_ref_field('monitor_id')),
            'monitor_ref_monitor_name': ('monitor_ref.monitor_name', _monitor_ref_field('monitor_name')),
            'monitor_ref_topic': ('monitor_ref.topic', _monitor_ref_field('topic')),
        }
    )

    # Карточка события
    columns.update(
        {
            'case_description': ('case.description', _case_field('description')),
            'case_note': ('case.note', _case_field('note')),
            'case_created_at': ('case.created_at', _case_field('created_at')),
            'case_updated_at': ('case.updated_at', _case_field('updated_at')),
            'case_created_by': ('case.created_by', _case_user('created_by')),
            'case_updated_by': ('case.updated_by', _case_user('updated_by')),
        }
    )

    # Документы
    columns.update(
        {
            'documents_count': ('documents_count', _docs_count),
            'documents_titles': ('documents_titles', _docs_titles),
            'documents_urls': ('documents_urls', _docs_urls),
        }
    )

    return columns

@login_required
@require_GET
def events_export_xlsx(request: HttpRequest):
    _assert_events_access(request.user)
    user = request.user
    allowed_monitors = _allowed_monitors_qs(user)
    form = AlarmFilterForm(request.GET or None, allowed_monitors_qs=allowed_monitors)
    if not form.is_valid():
        raise Http404('Invalid filter')

    # Сортировка экспорта — повторяем логику сортировки таблицы.
    sort = str(request.GET.get('sort') or 'time')
    direction = str(request.GET.get('dir') or 'desc').lower()
    allowed_sorts = {'time', 'monitor', 'topic', 'plate', 'description'}
    if sort not in allowed_sorts:
        sort = 'time'
    if direction not in ('asc', 'desc'):
        direction = 'desc'

    fields_req = request.GET.getlist('fields')
    columns = _export_columns(request)
    field_keys = [k for k in fields_req if k in columns]
    # Если пользователь не выбирал поля явно — экспортируем все (как требуется).
    if not field_keys:
        field_keys = list(sorted(columns.keys()))

    alarms = _filter_alarms(user=user, cleaned=form.cleaned_data, sort=sort, direction=direction)

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise Http404(f'openpyxl is required: {e}')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Events'

    headers = [columns[k][0] for k in field_keys]
    ws.append(headers)

    # --- Formatting: header row ---
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 22
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='000000')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

        # Ширина колонок — эвристика: от заголовка + небольшие поправки для "тяжёлых" полей.
        t = str(title or '')
        tl = t.lower()
        width = max(12, min(45, len(t) + 2))
        if 'url' in tl:
            width = 45
        elif 'identities' in tl or 'snapshots' in tl or 'data' == tl:
            width = 35
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for a in alarms:
        ws.append([columns[k][1](a) for k in field_keys])

    # Включаем автофильтр по диапазону данных (удобно для Excel)
    try:
        ws.auto_filter.ref = ws.dimensions
    except Exception:
        pass

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"events_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp

@login_required
@require_GET
def event_export_xlsx(request: HttpRequest, alarm_pk: int):
    _assert_events_access(request.user)
    # Загружаем Alarm с select_related (для FK)
    alarm = get_object_or_404(
        Alarm.objects.select_related('monitor_ref', 'account'),
        pk=int(alarm_pk)
    )
    if not _user_can_access_alarm(user=request.user, alarm=alarm):
        raise Http404('Forbidden')

    # Загружаем AlarmCase и AlarmDocument явными запросами (не через prefetch_related,
    # т.к. reverse OneToOneField + prefetch на единичном объекте ненадёжен).
    case = AlarmCase.objects.filter(alarm=alarm).select_related('created_by', 'updated_by').first()
    docs = list(case.documents.all().order_by('-uploaded_at')) if case else []

    # Привязываем case к alarm, чтобы _export_columns тоже видели данные
    if case is not None:
        alarm.case = case

    # DEBUG: выводим в логи информацию о найденных документах
    logger.info(
        "event_export_xlsx: alarm_pk=%s alarm_id=%s case_id=%s docs_count=%s files=%s",
        alarm.pk,
        getattr(alarm, 'alarm_id', ''),
        getattr(case, 'pk', None) if case else None,
        len(docs),
        [str(getattr(d.file, 'name', '') or '') for d in docs],
    )

    # Получаем скриншот события
    snap_bytes = _fetch_alarm_snapshot_bytes(alarm)

    try:
        import openpyxl
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise Http404(f'openpyxl is required: {e}')

    columns = _export_columns(request)
    keys = list(sorted(columns.keys()))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Event'

    # Заголовки
    ws.append([columns[k][0] for k in keys])
    # Значения (одна строка)
    ws.append([columns[k][1](alarm) for k in keys])
    ws.freeze_panes = 'A2'

    # Стилизация заголовков
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='4472C4')
    for col_idx in range(1, len(keys) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # Дополнительно: удобный для чтения формат key/value (как раньше), чтобы
    # «подробные поля» было проще смотреть в Excel.
    ws_kv = wb.create_sheet('Event (KV)')
    ws_kv.append(['field', 'value'])
    for k in keys:
        ws_kv.append([columns[k][0], columns[k][1](alarm)])
    ws_kv.freeze_panes = 'A2'
    ws_kv.column_dimensions['A'].width = 25
    ws_kv.column_dimensions['B'].width = 60

    # Стилизация KV заголовков
    for col_idx in range(1, 3):
        cell = ws_kv.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Лист со скриншотом
    if snap_bytes:
        ws_snap = wb.create_sheet('Screenshot')
        ws_snap.column_dimensions['A'].width = 60
        try:
            from PIL import Image as PILImage
            pil_img = PILImage.open(io.BytesIO(snap_bytes))
            # Масштабируем до разумного размера
            max_width = 600
            if pil_img.width > max_width:
                ratio = max_width / pil_img.width
                new_height = int(pil_img.height * ratio)
                pil_img = pil_img.resize((max_width, new_height), PILImage.LANCZOS)
            # Сохраняем во временный буфер
            img_buf = io.BytesIO()
            pil_img.save(img_buf, format='PNG')
            img_buf.seek(0)
            xl_img = XLImage(img_buf)
            ws_snap.add_image(xl_img, 'A1')
            ws_snap.row_dimensions[1].height = int(pil_img.height * 0.75)  # точки -> пункты
        except Exception as e:
            logger.warning("event_export_xlsx: failed to embed screenshot: %s", e)
            ws_snap['A1'] = 'Не удалось встроить изображение'

    # Лист с документами
    ws2 = wb.create_sheet('Documents')
    ws2.append(['title', 'file', 'uploaded_at', 'url'])
    for d in docs:
        url = request.build_absolute_uri(reverse('alarm_document_file', args=[d.pk]))
        ws2.append([
            str(d.title or ''),
            str(getattr(d.file, 'name', '') or ''),
            d.uploaded_at.strftime('%Y-%m-%d %H:%M:%S') if getattr(d, 'uploaded_at', None) else '',
            url,
        ])
    ws2.freeze_panes = 'A2'
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 40
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 50

    # Стилизация заголовков документов
    for col_idx in range(1, 5):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Лист с изображениями документов
    image_docs = []
    for d in docs:
        try:
            # Проверяем, что файл существует
            if not d.file or not hasattr(d.file, 'path'):
                logger.warning("event_export_xlsx: doc pk=%s has no file or path", d.pk)
                continue
            path = d.file.path
            import os
            if not os.path.exists(path):
                logger.warning("event_export_xlsx: doc pk=%s file path does not exist: %s", d.pk, path)
                continue
            ctype, _ = mimetypes.guess_type(path)
            logger.info("event_export_xlsx: doc pk=%s path=%s ctype=%s", d.pk, path, ctype)
            if ctype and ctype.startswith('image/'):
                image_docs.append((d, path))
        except Exception as e:
            logger.warning("event_export_xlsx: failed to check doc pk=%s: %s", d.pk, e)
            continue

    logger.info("event_export_xlsx: found %s image docs out of %s total docs", len(image_docs), len(docs))

    if image_docs:
        ws_imgs = wb.create_sheet('Document Images')
        ws_imgs.column_dimensions['A'].width = 80
        try:
            from PIL import Image as PILImage
            current_row = 1
            for d, path in image_docs:
                try:
                    name = str(d.title or '') or str(getattr(d.file, 'name', '') or '').split('/')[-1]
                    # Добавляем название документа
                    ws_imgs.cell(row=current_row, column=1, value=name)
                    ws_imgs.cell(row=current_row, column=1).font = Font(bold=True)
                    current_row += 1
                    
                    # Загружаем и масштабируем изображение
                    pil_img = PILImage.open(path)
                    max_width = 600
                    if pil_img.width > max_width:
                        ratio = max_width / pil_img.width
                        new_height = int(pil_img.height * ratio)
                        pil_img = pil_img.resize((max_width, new_height), PILImage.LANCZOS)
                    
                    # Сохраняем во временный буфер
                    img_buf = io.BytesIO()
                    pil_img.save(img_buf, format='PNG')
                    img_buf.seek(0)
                    xl_img = XLImage(img_buf)
                    ws_imgs.add_image(xl_img, f'A{current_row}')
                    
                    # Устанавливаем высоту строки
                    ws_imgs.row_dimensions[current_row].height = int(pil_img.height * 0.75)
                    current_row += 3  # Отступ между изображениями
                    logger.info("event_export_xlsx: successfully embedded image doc pk=%s", d.pk)
                except Exception as img_e:
                    logger.warning("event_export_xlsx: failed to embed image doc pk=%s: %s", d.pk, img_e)
                    continue
        except Exception as e:
            logger.warning("event_export_xlsx: failed to embed document images: %s", e)
            ws_imgs['A1'] = 'Не удалось встроить изображения документов'

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"event_{alarm.alarm_id or alarm.pk}.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _fetch_alarm_snapshot_bytes(alarm: Alarm) -> Optional[bytes]:
    snap = getattr(alarm, 'original_quality_snapshot', None)
    if not snap:
        return None
    try:
        account_id = int(alarm.account_id)
    except Exception:
        return None
    base_url = getattr(settings, 'BASE_URL', None)
    if not base_url:
        return None

    client = VAApiClient(account_id=account_id, base_url=str(base_url))
    try:
        resp = client.request('GET', str(snap), timeout=(7.0, 25.0))
        try:
            if resp.status_code != 200:
                return None
            return resp.content
        finally:
            resp.close()
    except Exception:
        return None

@login_required
@require_GET
def event_export_pdf(request: HttpRequest, alarm_pk: int):
    _assert_events_access(request.user)
    # Загружаем Alarm с select_related (для FK)
    alarm = get_object_or_404(
        Alarm.objects.select_related('monitor_ref', 'account'),
        pk=int(alarm_pk)
    )
    if not _user_can_access_alarm(user=request.user, alarm=alarm):
        raise Http404('Forbidden')

    # Загружаем AlarmCase и AlarmDocument явными запросами (не через prefetch_related,
    # т.к. reverse OneToOneField + prefetch на единичном объекте ненадёжен).
    case = AlarmCase.objects.filter(alarm=alarm).select_related('created_by', 'updated_by').first()
    docs = list(case.documents.all().order_by('-uploaded_at')) if case else []

    # Привязываем case к alarm, чтобы _export_columns тоже видели данные
    # (getattr(alarm, 'case', None) будет возвращать наш case)
    if case is not None:
        alarm.case = case

    # DEBUG: выводим в логи информацию о найденных документах
    try:
        logger.warning(
            "event_export_pdf: alarm_pk=%s alarm_id=%s case_id=%s docs=%s files=%s titles=%s",
            alarm.pk,
            getattr(alarm, 'alarm_id', ''),
            getattr(case, 'pk', None) if case else None,
            len(docs),
            [str(getattr(d.file, 'name', '') or '') for d in docs],
            [str(getattr(d, 'title', '') or '') for d in docs],
        )
    except Exception:
        logger.exception("event_export_pdf: failed to log docs")
    snap_bytes = _fetch_alarm_snapshot_bytes(alarm)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (  # type: ignore
            Image,
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
        from reportlab.lib import colors
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except Exception as e:
        raise Http404(f'reportlab is required for PDF export: {e}')

    # Шрифт с кириллицей
    font_name = 'Helvetica'
    try:
        dejavu_path = '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'
        pdfmetrics.registerFont(TTFont('DejaVuSans', dejavu_path))
        font_name = 'DejaVuSans'
    except Exception:
        pass

    bio = io.BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f'Alarm {alarm.alarm_id}',
    )

    styles = getSampleStyleSheet()
    style_title = ParagraphStyle(name='TitleDejaVu', parent=styles['Title'], fontName=font_name)
    style_h = ParagraphStyle(name='HDejaVu', parent=styles['Heading3'], fontName=font_name)
    style_n = ParagraphStyle(name='NDejaVu', parent=styles['BodyText'], fontName=font_name, fontSize=9, leading=11)

    title = f"Событие: {alarm.topic} / {alarm.monitor_name}"
    story = [Paragraph(title, style_title), Spacer(1, 4 * mm)]

    # Явно показываем количество документов рядом с шапкой (чтобы точно было на первой странице)
    story.append(Paragraph(f"Документы: {len(docs)}", style_n))
    story.append(Spacer(1, 2 * mm))

    # Таблица полей (используем Paragraph, чтобы переносить длинные строки)
    dt_start = _alarm_ts_to_aware_utc(alarm.start_time)
    dt_end = _alarm_ts_to_aware_utc(alarm.end_time)
    def P(x: str) -> Paragraph:
        return Paragraph(str(x or ''), style_n)

    rows = [
        [P('alarm_id'), P(str(alarm.alarm_id or ''))],
        [P('event_id'), P(str(alarm.event_id or ''))],
        [P('topic'), P(str(alarm.topic or ''))],
        [P('monitor_name'), P(str(alarm.monitor_name or ''))],
        [P('start_time'), P(timezone.localtime(dt_start).strftime('%Y-%m-%d %H:%M:%S') if dt_start else '')],
        [P('end_time'), P(timezone.localtime(dt_end).strftime('%Y-%m-%d %H:%M:%S') if dt_end else '')],
        [P('plate_numbers'), P(_plate_numbers_from_alarm(alarm))],
        [P('plate_identities'), P(_identities_to_text(getattr(alarm, 'plate_identities', None)))],
        [P('face_identities'), P(_identities_to_text(getattr(alarm, 'face_identities', None)))],
    ]
    if case:
        rows += [
            [P('case.description'), P(str(getattr(case, 'description', '') or ''))],
            [P('case.note'), P(str(getattr(case, 'note', '') or ''))],
        ]
    rows += [[P('documents_count'), P(str(len(docs)))]]

    t = Table(rows, colWidths=[45 * mm, None])
    t.setStyle(
        TableStyle(
            [
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOX', (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ('BACKGROUND', (0, 0), (0, -1), colors.whitesmoke),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(t)
    story.append(Spacer(1, 4 * mm))

    # Скриншот события
    story.append(Paragraph('Скриншот события', style_h))
    if snap_bytes:
        try:
            img = Image(io.BytesIO(snap_bytes))
            img.drawWidth = 160 * mm
            img.drawHeight = 90 * mm
            story.append(img)
        except Exception:
            story.append(Paragraph('Не удалось встроить изображение (доступно через ссылку в UI).', style_n))
    else:
        story.append(Paragraph('Скриншот недоступен.', style_n))
    story.append(Spacer(1, 4 * mm))

    # Документы
    # Часто документы не помещаются на 1 страницу после большого snapshot,
    # поэтому начинаем их с новой страницы.
    story.append(PageBreak())
    story.append(Paragraph(f'Документы ({len(docs)})', style_h))
    if not docs:
        story.append(Paragraph('Нет документов.', style_n))
    if docs:
        doc_rows = [[P('title'), P('file'), P('url')]]
        for d in docs:
            name = str(d.title or '') or str(getattr(d.file, 'name', '') or '').split('/')[-1]
            url = request.build_absolute_uri(reverse('alarm_document_file', args=[d.pk]))
            doc_rows.append([P(name), P(str(getattr(d.file, 'name', '') or '')), P(url)])

        dt = Table(doc_rows, colWidths=[55 * mm, 55 * mm, None])
        dt.setStyle(
            TableStyle(
                [
                    ('FONTNAME', (0, 0), (-1, -1), font_name),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('BOX', (0, 0), (-1, -1), 0.25, colors.lightgrey),
                    ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(dt)
        story.append(Spacer(1, 4 * mm))

        # Картинки документов — ниже таблицы
        story.append(Paragraph('Изображения документов (если есть)', style_h))
        image_count = 0
        for d in docs:
            try:
                # Проверяем, что файл существует
                if not d.file or not hasattr(d.file, 'path'):
                    logger.warning("event_export_pdf: doc pk=%s has no file or path", d.pk)
                    continue
                path = d.file.path
                import os
                if not os.path.exists(path):
                    logger.warning("event_export_pdf: doc pk=%s file path does not exist: %s", d.pk, path)
                    continue
                ctype, _ = mimetypes.guess_type(path)
                logger.info("event_export_pdf: doc pk=%s path=%s ctype=%s", d.pk, path, ctype)
                if ctype and ctype.startswith('image/'):
                    name = str(d.title or '') or str(getattr(d.file, 'name', '') or '').split('/')[-1]
                    story.append(Paragraph(name, style_n))
                    img = Image(path)
                    img.drawWidth = 160 * mm
                    img.drawHeight = 90 * mm
                    story.append(img)
                    story.append(Spacer(1, 3 * mm))
                    image_count += 1
            except Exception as e:
                logger.warning("event_export_pdf: failed to process doc pk=%s: %s", d.pk, e)
                continue
        logger.info("event_export_pdf: embedded %s images out of %s docs", image_count, len(docs))

    doc.build(story)
    pdf_bytes = bio.getvalue()
    filename = f"event_{alarm.alarm_id or alarm.pk}.pdf"
    resp = HttpResponse(pdf_bytes, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
